import streamlit as st
import pandas as pd
import numpy as np
import io
import sys, os; sys.path.insert(0, os.path.dirname(__file__))
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.worksheet.worksheet import Worksheet
from theme import apply_theme, page_header

st.set_page_config(page_title="Rapprochement Facture vs BL", page_icon="\u2696\ufe0f", layout="wide")
apply_theme()

# --- Constantes colonnes SOURCE (noms canoniques attendus) ---
COL_BL_REF_FOURN  = "R\u00e9f\u00e9rence Fournisseur"
COL_BL_DES        = "D\u00e9signation"
COL_BL_PU         = "P.U. HT"
COL_BL_QTE        = "Quantit\u00e9"
COL_BL_REM        = "Remise"
COL_BL_MNT        = "Montant HT"
COL_BL_NUM        = "N\u00b0 de pi\u00e8ce"

COL_FAC_REF_FOURN = "Af_RefFourniss"
COL_FAC_PU        = "Prix unitaire"
COL_FAC_REM       = "Remise"
COL_FAC_QTE       = "Quantit\u00e9"
COL_FAC_MNT       = "Montant HT"

# --- Constantes colonnes RESULTAT (noms d affichage uniques) ---
RES_REF       = "R\u00e9f\u00e9rence"
RES_DES       = "D\u00e9signation"
RES_NUM       = "N\u00b0 BL"
RES_FAC_QTE   = "Qt\u00e9 Facture"
RES_FAC_PU    = "PU Facture"
RES_FAC_REM   = "Remise Facture (%)"
RES_FAC_MNT   = "Montant HT Facture"
RES_BL_QTE    = "Qt\u00e9 BL"
RES_BL_PU     = "PU BL"
RES_BL_REM    = "Remise BL (%)"
RES_BL_MNT    = "Montant HT BL"
RES_ECART_QTE  = "\u00c9cart Qt\u00e9"
RES_ECART_PRIX = "\u00c9cart Prix U."
RES_ECART_REM  = "\u00c9cart Remise (%)"
RES_ECART_MNT  = "\u00c9cart Montant HT"
RES_STATUT     = "Statut"

defaults = {
    "data_loaded": False,
    "df_bl_raw": pd.DataFrame(),
    "df_fac_raw": pd.DataFrame(),
    "df_resultat": pd.DataFrame(),
    "manual_map": {}
}
for k, v in defaults.items():
    if k not in st.session_state:
        st.session_state[k] = v

# --- Utilitaires ---
def clean_currency(series):
    return pd.to_numeric(
        series.astype(str).str.replace("\u20ac","").str.replace(" ","").str.replace(",","."),
        errors="coerce"
    ).fillna(0.0)

def clean_text_id(series):
    return series.astype(str).str.replace(r"\.0$", "", regex=True).str.strip()

def nettoyer_ref(series):
    return series.astype(str).str.strip().str.upper()

def normaliser_colonnes(df, cibles):
    """Renomme les colonnes du df vers leur forme canonique (insensible a la casse)."""
    col_lower = {c.lower(): c for c in df.columns}
    rename = {}
    for cible in cibles:
        if cible not in df.columns and cible.lower() in col_lower:
            rename[col_lower[cible.lower()]] = cible
    return df.rename(columns=rename) if rename else df

# --- Styling Excel ---
def apply_excel_styling(ws: Worksheet, df_columns: list):
    header_fill = PatternFill(start_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    error_fill  = PatternFill(start_color="FFC7CE", fill_type="solid")
    error_font  = Font(color="9C0006")
    ok_font     = Font(color="006100")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"),  bottom=Side(style="thin"))
    euro_format = "#,##0.00 \u20ac"
    qty_format  = "0.00"
    cols_euro = [RES_FAC_PU, RES_FAC_MNT, RES_BL_PU, RES_BL_MNT, RES_ECART_PRIX, RES_ECART_MNT]
    cols_qty  = [RES_FAC_QTE, RES_BL_QTE, RES_ECART_QTE, RES_FAC_REM, RES_BL_REM, RES_ECART_REM]
    for row_idx, row in enumerate(ws.iter_rows(), 1):
        for cell in row:
            col_name = df_columns[cell.column - 1]
            cell.border = thin_border
            if row_idx == 1:
                cell.fill = header_fill; cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            else:
                if col_name in cols_euro: cell.number_format = euro_format
                elif col_name in cols_qty: cell.number_format = qty_format
                elif col_name == RES_NUM:
                    cell.number_format = "@"; cell.alignment = Alignment(horizontal="left")
                if col_name in [RES_ECART_QTE, RES_ECART_PRIX, RES_ECART_MNT, RES_ECART_REM]:
                    val = cell.value if isinstance(cell.value, (int, float)) else 0
                    if abs(val) > 0.01: cell.fill = error_fill; cell.font = error_font
                    else: cell.font = ok_font
    for col in ws.columns:
        ml = max((len(str(cell.value)) for cell in col if cell.value), default=0)
        ws.column_dimensions[col[0].column_letter].width = min((ml + 2) * 1.2, 50)

def style_recap_sheet(ws: Worksheet):
    bold_font = Font(bold=True); euro_format = "#,##0.00 \u20ac"
    ws.column_dimensions["A"].width = 50; ws.column_dimensions["B"].width = 25
    for row in ws.iter_rows():
        row[0].font = bold_font
        if row[0].row <= 3: row[1].number_format = euro_format; row[1].font = bold_font

def generer_excel(df):
    output = io.BytesIO()
    tot_fac = df[RES_FAC_MNT].sum(); tot_bl = df[RES_BL_MNT].sum()
    mask_prix      = abs(df[RES_ECART_PRIX]) > 0.01
    mask_rem       = abs(df[RES_ECART_REM])  > 0.01
    mask_manque_bl = df[RES_STATUT] == "Manque BL"
    mask_non_fac   = df[RES_STATUT] == "Non Factur\u00e9"
    mask_ecarts    = df[RES_STATUT] == "\u00c9cart Prix/Remise/Qt\u00e9"
    data_recap = [
        ("Montant total de la facture", tot_fac),
        ("Montant total des BL", tot_bl),
        ("\u00c9cart Global", tot_fac - tot_bl), ("",""),
        ("Nb produits avec \u00e9carts", mask_ecarts.sum()),
        ("Nb \u00e9carts de prix", mask_prix.sum()),
        ("Nb \u00e9carts de remises", mask_rem.sum()),
        ("Nb factur\u00e9s absents du BL", mask_manque_bl.sum()),
        ("Nb BL non factur\u00e9s", mask_non_fac.sum()),
    ]
    df_recap = pd.DataFrame(data_recap, columns=["Indicateur","Valeur"])
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_recap.to_excel(writer, sheet_name="R\u00e9capitulatif", index=False, header=False)
        style_recap_sheet(writer.sheets["R\u00e9capitulatif"])
        df.to_excel(writer, sheet_name="Rapprochement Global", index=False)
        apply_excel_styling(writer.sheets["Rapprochement Global"], df.columns.tolist())
        for mask, name in [(mask_prix,"Ecarts Prix"),(mask_rem,"Ecarts Remise"),
                           (mask_manque_bl,"Absents BL"),(mask_non_fac,"Non Factur\u00e9s")]:
            if mask.sum() > 0:
                df[mask].to_excel(writer, sheet_name=name, index=False)
                apply_excel_styling(writer.sheets[name], df.columns.tolist())
    output.seek(0); return output

# --- Interface ---
page_header("⚖️ Rapprochement Facture Fournisseur vs BL", "Comparez votre facture fournisseur avec vos bons de livraison et identifiez les écarts de prix, remise ou quantité.")

with st.sidebar:
    st.header("1. Importation")
    if st.button("\U0001f504\ufe0f R\u00e9initialiser tout"):
        st.session_state.clear(); st.rerun()
    mode = st.radio("Source des donn\u00e9es", ["2 Fichiers distincts", "1 Fichier (onglets)"])
    file_bl, file_fac, sheet_bl, sheet_fac = None, None, 0, 0
    if mode == "2 Fichiers distincts":
        file_fac = st.file_uploader("Fichier FACTURE", type=["xlsx"])
        file_bl  = st.file_uploader("Fichier BL",      type=["xlsx"])
    else:
        file_unique = st.file_uploader("Fichier Excel complet", type=["xlsx"])
        if file_unique:
            xl = pd.ExcelFile(file_unique)
            sheet_fac = st.selectbox("Onglet FACTURE", xl.sheet_names, index=0)
            sheet_bl  = st.selectbox("Onglet BL", xl.sheet_names,
                                     index=1 if len(xl.sheet_names) > 1 else 0)
            file_fac = file_unique; file_bl = file_unique

    if file_fac and file_bl:
        if st.button("Charger les donn\u00e9es", type="primary"):
            try:
                df_fac = pd.read_excel(file_fac, sheet_name=sheet_fac, header=0)
                df_bl  = pd.read_excel(file_bl,  sheet_name=sheet_bl,  header=0)

                # Normalisation de la casse des colonnes
                df_bl  = normaliser_colonnes(df_bl,  [COL_BL_REF_FOURN, COL_BL_DES, COL_BL_PU,
                                                       COL_BL_QTE, COL_BL_REM, COL_BL_MNT, COL_BL_NUM])
                df_fac = normaliser_colonnes(df_fac, [COL_FAC_REF_FOURN, COL_FAC_PU,
                                                       COL_FAC_QTE, COL_FAC_REM, COL_FAC_MNT])

                if COL_BL_REF_FOURN in df_bl.columns and COL_FAC_REF_FOURN in df_fac.columns:
                    for col in [COL_BL_PU, COL_BL_QTE, COL_BL_REM, COL_BL_MNT]:
                        if col in df_bl.columns: df_bl[col] = clean_currency(df_bl[col])
                    if COL_BL_NUM in df_bl.columns:
                        df_bl[COL_BL_NUM] = clean_text_id(df_bl[COL_BL_NUM])
                    for col in [COL_FAC_PU, COL_FAC_QTE, COL_FAC_REM, COL_FAC_MNT]:
                        if col in df_fac.columns: df_fac[col] = clean_currency(df_fac[col])
                    # Normalisation remise BL : d\u00e9cimal (0.40) \u2192 pourcentage (40)
                    if COL_BL_REM in df_bl.columns:
                        if df_bl[COL_BL_REM].dropna().abs().max() <= 1.0:
                            df_bl[COL_BL_REM] = df_bl[COL_BL_REM] * 100
                    df_bl.insert(0, "S\u00e9lectionner", True)
                    st.session_state.df_bl_raw  = df_bl
                    st.session_state.df_fac_raw = df_fac
                    st.session_state.manual_map = {}
                    st.session_state.data_loaded = True
                    st.rerun()
                else:
                    missing = []
                    if COL_BL_REF_FOURN not in df_bl.columns:
                        missing.append(f"'{COL_BL_REF_FOURN}' dans BL (trouv\u00e9es : {df_bl.columns.tolist()})")
                    if COL_FAC_REF_FOURN not in df_fac.columns:
                        missing.append(f"'{COL_FAC_REF_FOURN}' dans Facture (trouv\u00e9es : {df_fac.columns.tolist()})")
                    st.error("Colonnes cl\u00e9s manquantes :\n" + "\n".join(missing))
            except Exception as e:
                st.error(f"Erreur lecture : {e}")

if st.session_state.data_loaded:
    tab1, tab2, tab3 = st.tabs(["\u2705 S\u00e9lection BL", "\U0001f517 Rapprochement Manuel", "\U0001f4ca R\u00e9sultats & Export"])

    with tab1:
        st.info("D\u00e9cochez les lignes de BL \u00e0 ignorer.")
        cols_view  = ["S\u00e9lectionner", COL_BL_NUM, COL_BL_REF_FOURN, COL_BL_DES, COL_BL_PU, COL_BL_REM, COL_BL_QTE]
        cols_exist = [c for c in cols_view if c in st.session_state.df_bl_raw.columns]
        edited_bl  = st.data_editor(
            st.session_state.df_bl_raw[cols_exist],
            column_config={"S\u00e9lectionner": st.column_config.CheckboxColumn(default=True)},
            disabled=[c for c in cols_exist if c != "S\u00e9lectionner"],
            hide_index=True, use_container_width=True, height=400, key="editor_bl"
        )

    bl_selected_indices = edited_bl[edited_bl["S\u00e9lectionner"] == True].index
    df_bl_active = st.session_state.df_bl_raw.loc[bl_selected_indices].copy()

    with tab2:
        st.markdown("#### \U0001f517 Lier des r\u00e9f\u00e9rences orphelines")
        df_bl_mapped = df_bl_active.copy()
        df_bl_mapped["Ref_Join"] = nettoyer_ref(df_bl_mapped[COL_BL_REF_FOURN])
        for ref_bl_orig, ref_fac_target in st.session_state.manual_map.items():
            mask = df_bl_mapped["Ref_Join"] == nettoyer_ref(pd.Series([ref_bl_orig]))[0]
            df_bl_mapped.loc[mask, "Ref_Join"] = nettoyer_ref(pd.Series([ref_fac_target]))[0]
        refs_bl  = set(df_bl_mapped["Ref_Join"].unique())
        df_fac_c = st.session_state.df_fac_raw.copy()
        df_fac_c["Ref_Join"] = nettoyer_ref(df_fac_c[COL_FAC_REF_FOURN])
        refs_fac = set(df_fac_c["Ref_Join"].unique())
        orphans_fac = sorted(refs_fac - refs_bl)
        orphans_bl  = sorted(refs_bl  - refs_fac)
        c1, c2, c3 = st.columns([2, 2, 1])
        with c1:
            sel_fac = st.selectbox("1. R\u00e9f Facture orpheline", orphans_fac, key="sel_orph_fac") if orphans_fac else None
        with c2:
            if orphans_bl:
                def get_des(ref):
                    rows = df_bl_active[nettoyer_ref(df_bl_active[COL_BL_REF_FOURN]) == ref]
                    return str(rows.iloc[0].get(COL_BL_DES, ""))[:30] if not rows.empty else ""
                opts_bl = [f"{r} | {get_des(r)}" for r in orphans_bl]
                sel_bl_display = st.selectbox("2. R\u00e9f BL correspondante", opts_bl, key="sel_orph_bl")
                sel_bl = sel_bl_display.split(" | ")[0] if sel_bl_display else None
            else:
                sel_bl = None
        with c3:
            st.write("Action")
            if st.button("\u2795 Cr\u00e9er le lien") and sel_fac and sel_bl:
                st.session_state.manual_map[sel_bl] = sel_fac
                st.success(f"Lien cr\u00e9\u00e9 : {sel_bl} \u27a1\ufe0f {sel_fac}"); st.rerun()
        if st.session_state.manual_map:
            st.divider(); st.write("**Liens actifs :**"); to_del = None
            for k, v in st.session_state.manual_map.items():
                ca, cb = st.columns([4, 1]); ca.text(f"BL [{k}] fusionn\u00e9 vers [{v}]")
                if cb.button("\U0001f5d1\ufe0f", key=f"del_{k}"): to_del = k
            if to_del: del st.session_state.manual_map[to_del]; st.rerun()

    with tab3:
        if st.button("\U0001f680 Lancer le calcul", type="primary"):
            BL_QTE="_bl_qte"; BL_MNT="_bl_mnt"; BL_PU="_bl_pu"; BL_REM="_bl_rem"; BL_REF="_bl_ref"
            FAC_QTE="_fac_qte"; FAC_MNT="_fac_mnt"; FAC_PU="_fac_pu"; FAC_REM="_fac_rem"; FAC_REF="_fac_ref"

            df_bl_calc = df_bl_active.copy()
            df_bl_calc["_key_orig"] = nettoyer_ref(df_bl_calc[COL_BL_REF_FOURN])
            df_bl_calc["_key"] = df_bl_calc["_key_orig"]
            for ref_bl, ref_fac in st.session_state.manual_map.items():
                c_bl  = nettoyer_ref(pd.Series([ref_bl]))[0]
                c_fac = nettoyer_ref(pd.Series([ref_fac]))[0]
                df_bl_calc.loc[df_bl_calc["_key_orig"] == c_bl, "_key"] = c_fac
            df_bl_calc = df_bl_calc.rename(columns={
                COL_BL_REF_FOURN: BL_REF, COL_BL_QTE: BL_QTE,
                COL_BL_MNT: BL_MNT, COL_BL_PU: BL_PU, COL_BL_REM: BL_REM})
            agg_bl = {BL_REF:"first", BL_QTE:"sum", BL_MNT:"sum",
                      BL_PU:"first", BL_REM:"first", COL_BL_DES:"first"}
            if COL_BL_NUM in df_bl_calc.columns:
                agg_bl[COL_BL_NUM] = lambda x: ", ".join(sorted(set(x.astype(str))))
            grp_bl = df_bl_calc.groupby("_key").agg(agg_bl).reset_index()

            df_fac_calc = st.session_state.df_fac_raw.copy()
            df_fac_calc["_key"] = nettoyer_ref(df_fac_calc[COL_FAC_REF_FOURN])
            df_fac_calc = df_fac_calc.rename(columns={
                COL_FAC_REF_FOURN: FAC_REF, COL_FAC_QTE: FAC_QTE,
                COL_FAC_MNT: FAC_MNT, COL_FAC_PU: FAC_PU, COL_FAC_REM: FAC_REM})
            grp_fac = df_fac_calc.groupby("_key").agg({
                FAC_REF:"first", FAC_QTE:"sum", FAC_MNT:"sum",
                FAC_PU:"first", FAC_REM:"first"}).reset_index()

            merged = pd.merge(grp_fac, grp_bl, on="_key", how="outer", indicator=True)
            for c in [FAC_QTE, BL_QTE, FAC_MNT, BL_MNT, FAC_PU, BL_PU, FAC_REM, BL_REM]:
                merged[c] = merged[c].fillna(0)
            merged[RES_ECART_QTE]  = merged[FAC_QTE] - merged[BL_QTE]
            merged[RES_ECART_MNT]  = merged[FAC_MNT] - merged[BL_MNT]
            merged[RES_ECART_PRIX] = merged[FAC_PU]  - merged[BL_PU]
            merged[RES_ECART_REM]  = merged[FAC_REM] - merged[BL_REM]
            conditions = [
                merged["_merge"] == "left_only",
                merged["_merge"] == "right_only",
                (abs(merged[RES_ECART_MNT]) > 0.05) | (abs(merged[RES_ECART_REM]) > 0.01)
            ]
            merged[RES_STATUT] = np.select(conditions,
                ["Manque BL", "Non Factur\u00e9", "\u00c9cart Prix/Remise/Qt\u00e9"], default="OK")
            merged[RES_REF] = merged[FAC_REF].fillna(merged[BL_REF] + " (BL)")
            merged = merged.rename(columns={
                FAC_REF: "R\u00e9f Facture", BL_REF: "R\u00e9f BL",
                FAC_QTE: RES_FAC_QTE, BL_QTE: RES_BL_QTE,
                FAC_MNT: RES_FAC_MNT, BL_MNT: RES_BL_MNT,
                FAC_PU:  RES_FAC_PU,  BL_PU:  RES_BL_PU,
                FAC_REM: RES_FAC_REM, BL_REM: RES_BL_REM,
                COL_BL_DES: RES_DES,
            })
            if COL_BL_NUM in merged.columns:
                merged = merged.rename(columns={COL_BL_NUM: RES_NUM})
            final_cols = [c for c in [
                RES_REF, RES_DES, RES_NUM,
                RES_FAC_QTE, RES_FAC_PU, RES_FAC_REM, RES_FAC_MNT,
                RES_BL_QTE,  RES_BL_PU,  RES_BL_REM,  RES_BL_MNT,
                RES_ECART_QTE, RES_ECART_PRIX, RES_ECART_REM, RES_ECART_MNT,
                RES_STATUT
            ] if c in merged.columns]
            st.session_state.df_resultat = merged[final_cols].copy()

        if not st.session_state.df_resultat.empty:
            df_res = st.session_state.df_resultat
            c1, c2, c3 = st.columns(3)
            c1.metric("Total Facture", f"{df_res[RES_FAC_MNT].sum():,.2f} \u20ac")
            c2.metric("Total BL",      f"{df_res[RES_BL_MNT].sum():,.2f} \u20ac")
            c3.metric("\u00c9cart Global",
                      f"{(df_res[RES_FAC_MNT].sum()-df_res[RES_BL_MNT].sum()):,.2f} \u20ac",
                      delta_color="inverse")
            def highlight_errors(val):
                return "background-color: #ffcccc; color: #990000" \
                    if isinstance(val, (int, float)) and abs(val) > 0.01 else ""
            st.dataframe(
                df_res.style
                      .map(highlight_errors,
                           subset=[RES_ECART_QTE, RES_ECART_PRIX, RES_ECART_REM, RES_ECART_MNT])
                      .format({RES_BL_REM:"{:.2f}", RES_FAC_REM:"{:.2f}", RES_ECART_REM:"{:.2f}"}),
                use_container_width=True, height=500
            )
            fname = st.text_input("Nom du fichier", "Resultat_Rapprochement")
            st.download_button(
                "\U0001f4e5 T\u00e9l\u00e9charger Excel complet",
                data=generer_excel(df_res),
                file_name=f"{fname}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
