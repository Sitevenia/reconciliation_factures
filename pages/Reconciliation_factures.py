import streamlit as st
import pandas as pd
import numpy as np
import io
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.worksheet.worksheet import Worksheet

# --- Configuration de la page ---
st.set_page_config(page_title="Rapprochement Facture vs BL", page_icon="⚖️", layout="wide")

# --- Constantes des Colonnes (Basées sur vos captures) ---
# ONGLETS / FICHIERS BL
COL_BL_DATE = 'Date Document'
COL_BL_NUM = 'N° de pièce'
COL_BL_REF_ART = 'Référence Article'
COL_BL_REF_FOURN = 'AF_RefFourniss' # Clé de jointure
COL_BL_DES = 'Désignation Article'
COL_BL_PU = 'Prix Unitaire HT'
COL_BL_QTE = 'Qté Livrées'
COL_BL_REM = 'Pourcentage Remise'
COL_BL_MNT = 'Montant HT Net'

# ONGLET / FICHIER FACTURE
COL_FAC_REF_FOURN = 'Af_RefFourniss' # Clé de jointure
COL_FAC_PU = 'Prix unitaire'
COL_FAC_REM = 'Remise'
COL_FAC_QTE = 'Quantité'
COL_FAC_MNT = 'Montant HT'

# COLONNES CALCULÉES
COL_ECART_QTE = 'Écart Qté'
COL_ECART_PRIX = 'Écart Prix U.'
COL_ECART_MNT = 'Écart Montant HT'
COL_STATUT = 'Statut'

# --- Initialisation de l'État de Session ---
if 'data_loaded' not in st.session_state:
    st.session_state.data_loaded = False
if 'df_bl_raw' not in st.session_state:
    st.session_state.df_bl_raw = pd.DataFrame()
if 'df_fac_raw' not in st.session_state:
    st.session_state.df_fac_raw = pd.DataFrame()
if 'df_resultat' not in st.session_state:
    st.session_state.df_resultat = pd.DataFrame()

# --- Fonctions Utilitaires ---

def clean_currency(series):
    """Convertit en float les montants (gère , et €)."""
    return pd.to_numeric(
        series.astype(str).str.replace('€', '').str.replace(' ', '').str.replace(',', '.'), 
        errors='coerce'
    ).fillna(0.0)

def nettoyer_ref(series):
    """Nettoie les références pour la jointure (supprime espaces, met en majuscules)."""
    return series.astype(str).str.strip().str.upper()

# --- Fonction Export Excel (Adaptée de votre ancien code) ---
def apply_excel_styling_rapprochement(ws: Worksheet, df_columns: list):
    header_fill = PatternFill(start_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Styles pour les écarts
    error_fill = PatternFill(start_color="FFC7CE", fill_type="solid") # Rouge clair
    error_font = Font(color="9C0006") # Rouge foncé
    ok_fill = PatternFill(start_color="C6EFCE", fill_type="solid") # Vert clair
    ok_font = Font(color="006100") # Vert foncé
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    euro_format = '#,##0.00 €'
    qty_format = '0.00'
    
    # Mapping des formats
    cols_euro = [COL_BL_PU, COL_BL_MNT, COL_FAC_PU, COL_FAC_MNT, COL_ECART_PRIX, COL_ECART_MNT]
    cols_qty = [COL_BL_QTE, COL_FAC_QTE, COL_ECART_QTE, COL_BL_REM, COL_FAC_REM]

    for row_idx, row in enumerate(ws.iter_rows(), 1):
        for cell in row:
            col_name = df_columns[cell.column - 1]
            cell.border = thin_border
            
            # Header
            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            else:
                # Formatage Numérique
                if col_name in cols_euro:
                    cell.number_format = euro_format
                elif col_name in cols_qty:
                    cell.number_format = qty_format
                
                # Mise en évidence des écarts
                if col_name in [COL_ECART_QTE, COL_ECART_PRIX, COL_ECART_MNT]:
                    val = cell.value if isinstance(cell.value, (int, float)) else 0
                    if abs(val) > 0.01:
                        cell.fill = error_fill
                        cell.font = error_font
                    else:
                        cell.font = ok_font

    # Ajustement largeur colonnes
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = min(adjusted_width, 50)

from openpyxl.styles import Alignment # Import manquant ajouté

def generer_excel(df):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Rapprochement', index=False)
        apply_excel_styling_rapprochement(writer.sheets['Rapprochement'], df.columns.tolist())
    output.seek(0)
    return output

# --- Interface Principale ---
st.title("⚖️ Rapprochement Facture Fournisseur vs BL")

with st.sidebar:
    st.header("1. Importation")
    if st.button("🔄️ Réinitialiser"):
        st.session_state.clear()
        st.rerun()
        
    mode = st.radio("Source des données", ["2 Fichiers distincts", "1 Fichier (onglets)"])
    
    file_bl = None
    file_fac = None
    sheet_bl = 0
    sheet_fac = 0
    
    if mode == "2 Fichiers distincts":
        file_fac = st.file_uploader("Fichier FACTURE", type=["xlsx"])
        file_bl = st.file_uploader("Fichier BL", type=["xlsx"])
    else:
        file_unique = st.file_uploader("Fichier Excel complet", type=["xlsx"])
        if file_unique:
            xl = pd.ExcelFile(file_unique)
            sheet_fac = st.selectbox("Onglet FACTURE", xl.sheet_names, index=0)
            sheet_bl = st.selectbox("Onglet BL", xl.sheet_names, index=1 if len(xl.sheet_names)>1 else 0)
            file_fac = file_unique
            file_bl = file_unique

    # Bouton de chargement
    if file_fac and file_bl:
        if st.button("Charger les données", type="primary"):
            try:
                # Lecture FACTURE (Supposé header ligne 1 -> header=0)
                df_fac = pd.read_excel(file_fac, sheet_name=sheet_fac)
                
                # Lecture BL (Header ligne 3 -> header=2)
                df_bl = pd.read_excel(file_bl, sheet_name=sheet_bl, header=2)
                
                # Nettoyage préliminaire
                # On s'assure que les colonnes clés existent
                if COL_BL_REF_FOURN not in df_bl.columns:
                    st.error(f"Colonne '{COL_BL_REF_FOURN}' introuvable dans le BL (Ligne 3).")
                elif COL_FAC_REF_FOURN not in df_fac.columns:
                    st.error(f"Colonne '{COL_FAC_REF_FOURN}' introuvable dans la Facture.")
                else:
                    # Conversion des colonnes numériques BL
                    for col in [COL_BL_PU, COL_BL_QTE, COL_BL_REM, COL_BL_MNT]:
                        if col in df_bl.columns:
                            df_bl[col] = clean_currency(df_bl[col])
                    
                    # Conversion des colonnes numériques FACTURE
                    for col in [COL_FAC_PU, COL_FAC_QTE, COL_FAC_REM, COL_FAC_MNT]:
                        if col in df_fac.columns:
                            df_fac[col] = clean_currency(df_fac[col])
                            
                    # Création de la colonne de sélection pour les BL
                    df_bl.insert(0, "Sélectionner", True)
                    
                    st.session_state.df_bl_raw = df_bl
                    st.session_state.df_fac_raw = df_fac
                    st.session_state.data_loaded = True
                    st.rerun()
                    
            except Exception as e:
                st.error(f"Erreur lors de la lecture des fichiers : {e}")

# --- Logique de l'application ---

if st.session_state.data_loaded:
    
    # ---------------------------------------------------------
    # ÉTAPE 2 : SÉLECTION DES BL
    # ---------------------------------------------------------
    st.header("2. Sélection des Bons de Livraison")
    st.info("Décochez les lignes de BL que vous ne souhaitez pas inclure dans le rapprochement (ex: articles non facturés ce mois-ci).")
    
    # Configuration de l'éditeur pour afficher les infos utiles
    cols_bl_view = ["Sélectionner", COL_BL_NUM, COL_BL_REF_FOURN, COL_BL_DES, COL_BL_PU, COL_BL_REM, COL_BL_QTE, COL_BL_MNT]
    # On filtre pour n'afficher que les colonnes qui existent vraiment
    cols_bl_exist = [c for c in cols_bl_view if c in st.session_state.df_bl_raw.columns]
    
    edited_bl = st.data_editor(
        st.session_state.df_bl_raw[cols_bl_exist],
        column_config={
            "Sélectionner": st.column_config.CheckboxColumn("Inclure?", help="Cocher pour comparer", default=True),
            COL_BL_MNT: st.column_config.NumberColumn(format="%.2f €"),
            COL_BL_PU: st.column_config.NumberColumn(format="%.2f €"),
        },
        disabled=[c for c in cols_bl_exist if c != "Sélectionner"],
        hide_index=True,
        use_container_width=True,
        height=300
    )
    
    # ---------------------------------------------------------
    # ÉTAPE 3 : CALCUL & COMPARAISON
    # ---------------------------------------------------------
    if st.button("🚀 Lancer le rapprochement", type="primary"):
        
        # 1. Filtrer les BL sélectionnés
        bl_selected = edited_bl[edited_bl["Sélectionner"] == True].copy()
        
        # 2. Agréger les BL par Référence Fournisseur (cas où plusieurs BL contiennent le même produit)
        # On somme les quantités et les montants. On fait une moyenne pondérée ou max pour le prix/remise (à simplifier)
        grp_bl = bl_selected.groupby(COL_BL_REF_FOURN).agg({
            COL_BL_QTE: 'sum',
            COL_BL_MNT: 'sum',
            COL_BL_PU: 'first', # On suppose le prix unitaire constant
            COL_BL_REM: 'first',
            COL_BL_DES: 'first',
            COL_BL_NUM: lambda x: ', '.join(x.astype(str).unique()) # Liste des BL concaténés
        }).reset_index()
        
        # Nettoyage des clés pour la jointure
        grp_bl['_key'] = nettoyer_ref(grp_bl[COL_BL_REF_FOURN])
        
        df_fac = st.session_state.df_fac_raw.copy()
        df_fac['_key'] = nettoyer_ref(df_fac[COL_FAC_REF_FOURN])
        
        # 3. Jointure (Outer Join pour voir les manquants des deux côtés)
        merged = pd.merge(
            df_fac, 
            grp_bl, 
            on='_key', 
            how='outer', 
            suffixes=('_FAC', '_BL'),
            indicator=True
        )
        
        # 4. Calcul des écarts
        # On remplit les NaN par 0 pour les calculs
        merged[COL_FAC_QTE] = merged[COL_FAC_QTE].fillna(0)
        merged[COL_BL_QTE] = merged[COL_BL_QTE].fillna(0)
        merged[COL_FAC_MNT] = merged[COL_FAC_MNT].fillna(0)
        merged[COL_BL_MNT] = merged[COL_BL_MNT].fillna(0)
        merged[COL_FAC_PU] = merged[COL_FAC_PU].fillna(0)
        merged[COL_BL_PU] = merged[COL_BL_PU].fillna(0)

        # Calculs
        merged[COL_ECART_QTE] = merged[COL_FAC_QTE] - merged[COL_BL_QTE]
        merged[COL_ECART_MNT] = merged[COL_FAC_MNT] - merged[COL_BL_MNT]
        merged[COL_ECART_PRIX] = merged[COL_FAC_PU] - merged[COL_BL_PU]
        
        # Détermination du statut
        conditions = [
            (merged['_merge'] == 'left_only'), # Présent Facture, absent BL
            (merged['_merge'] == 'right_only'), # Présent BL, absent Facture
            (abs(merged[COL_ECART_MNT]) > 0.05) # Ecart montant significatif
        ]
        choices = ['Manque BL', 'Non Facturé', 'Écart Prix/Qté']
        merged[COL_STATUT] = np.select(conditions, choices, default='OK')
        
        # 5. Mise en forme finale du tableau
        # On priorise l'affichage de la référence Facture, sinon celle du BL
        merged['Référence'] = merged[COL_FAC_REF_FOURN].fillna(merged[COL_BL_REF_FOURN])
        
        # Sélection et ordre des colonnes
        final_cols = [
            'Référence',
            COL_BL_DES, # Désignation (venant du BL)
            COL_BL_NUM, # N° BL
            # Partie Facture
            COL_FAC_QTE, COL_FAC_PU, COL_FAC_REM, COL_FAC_MNT,
            # Partie BL
            COL_BL_QTE, COL_BL_PU, COL_BL_REM, COL_BL_MNT,
            # Ecarts
            COL_ECART_QTE, COL_ECART_PRIX, COL_ECART_MNT,
            COL_STATUT
        ]
        
        st.session_state.df_resultat = merged[final_cols].copy()

    # ---------------------------------------------------------
    # ÉTAPE 4 : RÉSULTATS
    # ---------------------------------------------------------
    if not st.session_state.df_resultat.empty:
        st.header("3. Analyse des Écarts")
        
        df_res = st.session_state.df_resultat
        
        # KPI Globaux
        tot_fac = df_res[COL_FAC_MNT].sum()
        tot_bl = df_res[COL_BL_MNT].sum()
        delta_tot = tot_fac - tot_bl
        
        c1, c2, c3 = st.columns(3)
        c1.metric("Total Facture HT", f"{tot_fac:,.2f} €")
        c2.metric("Total BL (Sélectionnés) HT", f"{tot_bl:,.2f} €")
        c3.metric("Écart Global", f"{delta_tot:,.2f} €", delta_color="inverse")
        
        # Affichage avec couleur conditionnelle simple pour Streamlit
        def highlight_vals(val):
            if isinstance(val, (int, float)) and abs(val) > 0.01:
                return 'color: red; font-weight: bold'
            return ''
        
        st.dataframe(
            df_res.style.applymap(highlight_vals, subset=[COL_ECART_QTE, COL_ECART_PRIX, COL_ECART_MNT]),
            use_container_width=True,
            height=500
        )
        
        # Export Excel
        st.write("---")
        st.header("4. Export")
        nom_fichier = st.text_input("Nom du fichier", "Resultat_Rapprochement")
        
        xlsx_data = generer_excel(df_res)
        
        st.download_button(
            label="📥 Télécharger le fichier Excel comparatif",
            data=xlsx_data,
            file_name=f"{nom_fichier}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary"
        )
